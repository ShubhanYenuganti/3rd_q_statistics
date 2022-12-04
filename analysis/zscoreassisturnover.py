from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import scipy.stats as stats

wb = load_workbook('/Users/shubhan/Desktop/Projects/3rd_q_statistics/3rd_Q_statistics.xlsx')
ws = wb.active
row = ws.max_row;

arr = []

for num in range(2, ws.max_row + 1):
    res = ws['' + get_column_letter(16)+str(num)].value / ws['' + get_column_letter(22)+str(num)].value
    arr.append(res)

data = np.array(arr)

# Z-Scores for Assists/Turnover Per Quarter
ws['' + get_column_letter(35)+str(1)] = "Z-Score for ATRPQ"
res = stats.zscore(data, axis = 0)

for num in range(2, ws.max_row + 1):
    ws['' + get_column_letter(35)+str(num)] = res[num - 2]

wb.save('/Users/shubhan/Desktop/Projects/3rd_q_statistics/3rd_Q_statistics.xlsx')