from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from results import getAllStats

wb = load_workbook('3rd_Q_statistics.xlsx')
ws = wb.active

row = ws.max_row + 1

# Appends data from results.py into a spreadsheet
for col in range(1, 28):
    char = get_column_letter(col)
    ws[char + str(row)] = getAllStats()[col - 1]

wb.save('3rd_Q_statistics.xlsx')